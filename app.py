import streamlit as st
import base64
from google import genai
from google.genai import types
from PIL import Image
import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import re
from typing import List, Dict
import tempfile
import os
import zipfile
import shutil


def extract_bold_key_value(line: str):
    match = re.search(r"\*\*(.+?)\*\*:\s*(.+)", line)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    return None, None

def parse_markdown_output(md_text: str):
    summary = {}
    standard = {}
    rows = []

    summary_section = re.search(r"### 1\. Test Summary Information(.*?)---", md_text, re.S)
    if summary_section:
        for line in summary_section.group(1).splitlines():
            key, val = extract_bold_key_value(line)
            if key:
                summary[key] = val

    std_section = re.search(r"### 2\. Verification with Standard Block(.*?)---", md_text, re.S)
    if std_section:
        for line in std_section.group(1).splitlines():
            key, val = extract_bold_key_value(line)
            if key:
                standard[key] = val

    table_match = re.search(r"\| Sr\. No\..*", md_text, re.S)
    if table_match:
        lines = table_match.group(0).splitlines()
        data_lines = [l for l in lines if l.startswith("|") and not "---" in l][1:]

        for line in data_lines:
            cols = [c.strip() for c in line.strip("|").split("|")]
            if len(cols) < 7:
                continue

            base_vals = [int(x.strip()) for x in cols[3].split(",") if x.strip().isdigit()]
            haz_vals = [int(x.strip()) for x in cols[4].split(",") if x.strip().isdigit()]
            weld_vals = [int(x.strip()) for x in cols[5].split(",") if x.strip().isdigit()]
            test_conducted = bool(base_vals or haz_vals or weld_vals)

            pipe_no = cols[1]
            if pipe_no.startswith('E'):
                pipe_no = 'N' + pipe_no[1:]

            rows.append({
                "sample_id": pipe_no,
                "heat_no": cols[2],
                "base": base_vals,
                "haz": haz_vals,
                "weld": weld_vals,
                "remarks": cols[6] if len(cols) > 6 else "",
                "test_conducted": test_conducted
            })

    return summary, standard, rows

def populate_excel_from_markdown(md_text, template_path, output_path):
    summary, standard, samples = parse_markdown_output(md_text)

    wb = load_workbook(template_path)
    ws = wb.active

    summary_map = {
        "Testing Laboratory": "B5", "Document Title": "B6", "Format No.": "B7",
        "Specification & Grade": "B8", "Test Method": "B9", "Pipe Size": "B12",
        "Atmospheric Conditions": "B13", "Date & Shift": "B14", "Requirements": "B15", "M/C No.": "B16",
    }
    for k, cell in summary_map.items():
        if k in summary:
            ws[cell] = summary[k]

    std_map = {
        "Standard Block ID No.": "E5", "Standard Block Value": "E6", "Reading 1": "E7",
        "Reading 2": "E8", "Reading 3": "E9", "Reading 4": "E10", "Reading 5": "E11",
        "Average (AVG)": "E12", "% Of Error": "E13", "Remark": "E14",
    }
    for k, cell in std_map.items():
        if k in standard:
            ws[cell] = standard[k]

    START_ROW = 17
    COL_SRNO, COL_HEAT, COL_SAMPLE = 1, 2, 3
    COL_BASE, COL_HAZ, COL_WELD, COL_REMARK = 4, 10, 28, 37

    for idx, s in enumerate(samples):
        r = START_ROW + idx
        ws.cell(r, COL_SRNO, idx + 1)
        ws.cell(r, COL_SAMPLE, s["sample_id"])
        ws.cell(r, COL_HEAT, s["heat_no"])
        for i, v in enumerate(s["base"]):
            ws.cell(r, COL_BASE + i, v)
        for i, v in enumerate(s["haz"]):
            ws.cell(r, COL_HAZ + i, v)
        for i, v in enumerate(s["weld"]):
            ws.cell(r, COL_WELD + i, v)
        ws.cell(r, COL_REMARK, s["remarks"])

    wb.save(output_path)

def create_validation_excel(samples: List[Dict], master_map: Dict, all_heats_in_master: List[str], output_path: str):
    """Create validation Excel file showing TESTED rows first, then NOT TESTED"""
    results_data = []
    
    # First, add all TESTED samples (with test_conducted = True)
    tested_heats = set()
    for sample in samples:
        if not sample.get("test_conducted", False):
            continue
            
        heat_no = sample.get('heat_no', '').strip()
        pipe_used = sample.get('sample_id', '').strip()
        tested_heats.add(heat_no)
        
        if heat_no and pipe_used:
            if heat_no in master_map:
                expected_pipe = master_map[heat_no]
                if pipe_used == expected_pipe:
                    results_data.append({
                        "Heat No": heat_no,
                        "Pipe No Used in Test": pipe_used,
                        "Status": "✅ TESTED - MATCHED",
                        "Result": f"Test conducted using Pipe No: {pipe_used}"
                    })
                else:
                    results_data.append({
                        "Heat No": heat_no,
                        "Pipe No Used in Test": pipe_used,
                        "Status": "⚠️ TESTED - MISMATCH",
                        "Result": f"Pipe mismatch! Test used {pipe_used}"
                    })
            else:
                results_data.append({
                    "Heat No": heat_no,
                    "Pipe No Used in Test": pipe_used,
                    "Status": "❌ TESTED - NOT IN MASTER",
                    "Result": f"Heat No {heat_no} not found in master list"
                })
    
    # Then, add all NOT TESTED heat numbers from master
    for heat_no in all_heats_in_master:
        if heat_no not in tested_heats:
            results_data.append({
                "Heat No": heat_no,
                "Pipe No Used in Test": "NOT TESTED",
                "Status": "❌ NOT TESTED",
                "Result": f"Heat No {heat_no} was not tested"
            })
    
    df = pd.DataFrame(results_data)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Validation Results', index=False)
        worksheet = writer.sheets['Validation Results']
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Add color coding
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        for row in worksheet.iter_rows(min_row=2, max_row=len(results_data) + 1):
            status_cell = row[2]  # Status column
            if "MATCHED" in str(status_cell.value):
                status_cell.fill = green_fill
            elif "NOT TESTED" in str(status_cell.value) or "NOT IN MASTER" in str(status_cell.value):
                status_cell.fill = red_fill
            elif "MISMATCH" in str(status_cell.value):
                status_cell.fill = yellow_fill


def validate_with_master_fast(samples: List[Dict], master_map: Dict, all_heats_in_master: List[str]) -> pd.DataFrame:
    """Fast validation showing TESTED first, then NOT TESTED"""
    results = []
    tested_heats = set()
    
    # First, add TESTED samples (only those with test_conducted = True)
    for sample in samples:
        if not sample.get("test_conducted", False):
            continue
        heat_no = sample.get('heat_no', '').strip()
        pipe_used = sample.get('sample_id', '').strip()
        
        if heat_no and pipe_used:
            tested_heats.add(heat_no)
            if heat_no in master_map:
                expected_pipe = master_map[heat_no]
                if pipe_used == expected_pipe:
                    results.append({"Heat No": heat_no, "Pipe No Used": pipe_used, "Status": "✅ TESTED - MATCHED"})
                else:
                    results.append({"Heat No": heat_no, "Pipe No Used": pipe_used, "Status": "⚠️ TESTED - MISMATCH"})
            else:
                results.append({"Heat No": heat_no, "Pipe No Used": pipe_used, "Status": "❌ TESTED - NOT IN MASTER"})
    
    # Then, add NOT TESTED heats from master
    for heat_no in all_heats_in_master:
        if heat_no not in tested_heats:
            results.append({"Heat No": heat_no, "Pipe No Used": "NOT TESTED", "Status": "❌ NOT TESTED"})
    
    return pd.DataFrame(results)

# -------------------------
# Initialize session state
# -------------------------
if 'extracted_text' not in st.session_state:
    st.session_state.extracted_text = None
if 'samples' not in st.session_state:
    st.session_state.samples = None
if 'master_map' not in st.session_state:
    st.session_state.master_map = None
if 'master_list_grouped' not in st.session_state:
    st.session_state.master_list_grouped = None
if 'all_heats_in_master' not in st.session_state:
    st.session_state.all_heats_in_master = None
if 'excel_ready' not in st.session_state:
    st.session_state.excel_ready = False
if 'validation_excel_ready' not in st.session_state:
    st.session_state.validation_excel_ready = False

# -------------------------
# Streamlit UI
# -------------------------
st.set_page_config(page_title="Vickers Hardness Sheet Extractor & Validator", layout="wide")
st.title("Vickers Hardness Observation Sheet Extractor & Validator")

# Sidebar for validation option
st.sidebar.header("Validation Options")
enable_validation = st.sidebar.checkbox("Enable Master - Observation Validation", value=False)

if enable_validation:
    st.sidebar.markdown("""
    **Upload Master Sheet Excel** containing:
    - Pipe No 
    - Heat No 
    
    This will validate tested samples against master list.
    """)
    
    master_file = st.sidebar.file_uploader("Upload Master Excel", type=["xlsx", "xls"], key="master_excel")
    
    if master_file and st.session_state.master_map is None:
        with st.sidebar.status("Loading master data from Excel..."):
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                    tmp_file.write(master_file.getvalue())
                    tmp_path = tmp_file.name
                
                # Try to read directly first
                try:
                    df_raw = pd.read_excel(tmp_path, header=None)
                except:
                    # Fix corrupted XML if needed
                    with zipfile.ZipFile(tmp_path, 'r') as zip_ref:
                        zip_ref.extractall("temp_excel")
                    
                    workbook_path = "temp_excel/xl/workbook.xml"
                    with open(workbook_path, "r", encoding="utf-8") as f:
                        content = f.read()
                    
                    content = re.sub(r"<definedNames>.*?</definedNames>", "", content, flags=re.DOTALL)
                    
                    with open(workbook_path, "w", encoding="utf-8") as f:
                        f.write(content)
                    
                    fixed_path = tmp_path.replace('.xlsx', '_fixed.xlsx')
                    shutil.make_archive(fixed_path.replace('.xlsx', ''), 'zip', "temp_excel")
                    os.rename(fixed_path.replace('.xlsx', '') + '.zip', fixed_path)
                    
                    df_raw = pd.read_excel(fixed_path, header=None)
                    os.unlink(fixed_path)
                    shutil.rmtree("temp_excel", ignore_errors=True)
                
                # Find header row with "Pipe No."
                header_row = df_raw[df_raw.eq("Pipe No.").any(axis=1)].index[0]
                
                df = df_raw.iloc[header_row:]
                df.columns = df.iloc[0]
                df = df[1:]
                
                result_df = df[["Pipe No.", "Heat No."]].dropna().reset_index(drop=True)
                
                master_map = {}
                heat_to_pipes = {}
                all_heats = []
                
                for _, row in result_df.iterrows():
                    pipe_no = str(row["Pipe No."]).strip()
                    heat_no_raw = str(row["Heat No."]).strip()
                    
                    if pipe_no and pipe_no != 'nan' and heat_no_raw and heat_no_raw != 'nan':
                        if '/' in heat_no_raw:
                            for h in heat_no_raw.split('/'):
                                h = h.strip()
                                master_map[h] = pipe_no
                                all_heats.append(h)
                                if h not in heat_to_pipes:
                                    heat_to_pipes[h] = []
                                if pipe_no not in heat_to_pipes[h]:
                                    heat_to_pipes[h].append(pipe_no)
                        else:
                            master_map[heat_no_raw] = pipe_no
                            all_heats.append(heat_no_raw)
                            if heat_no_raw not in heat_to_pipes:
                                heat_to_pipes[heat_no_raw] = []
                            if pipe_no not in heat_to_pipes[heat_no_raw]:
                                heat_to_pipes[heat_no_raw].append(pipe_no)
                
                os.unlink(tmp_path)
                
                grouped_data = []
                for heat_no, pipes in heat_to_pipes.items():
                    grouped_data.append({
                        "Heat No": heat_no,
                        "Associated Pipe Nos": "\n".join(sorted(pipes)),
                        "Number of Pipes": len(pipes)
                    })
                
                grouped_data.sort(key=lambda x: x["Heat No"])
                
                st.session_state.master_map = master_map
                st.session_state.master_list_grouped = pd.DataFrame(grouped_data)
                st.session_state.all_heats_in_master = list(set(all_heats))
                
                st.sidebar.success(f"✅ Loaded {len(master_map)} heat number mappings")
                st.sidebar.info(f"📊 {len(grouped_data)} unique Heat Numbers")
                
            except Exception as e:
                st.sidebar.error(f"Error loading Excel: {e}")

# Main content - Upload observation sheet
st.write("""
Upload an 'Observation Sheet (Mechanical - Vickers Hardness Test)' image and extract structured data.
""")

uploaded_file = st.file_uploader("Upload Observation Sheet Image", type=["png", "jpg", "jpeg", "tiff"])

if uploaded_file:
    image = Image.open(uploaded_file)
    st.image(image, caption="Uploaded Observation Sheet", use_container_width=True)

    buffered = io.BytesIO()
    image.save(buffered, format="JPEG")
    image_bytes = buffered.getvalue()

    prompt = """
Extract all information from the provided 'Observation Sheet (Mechanical - Vickers Hardness Test)' image into a structured Markdown format. 
Please organize the output into the following three sections:

1. Test Summary Information:
   Extract all metadata from the top section, including Customer, Spec & Grade, Test Method, Pipe Size, Atmospheric Conditions, Date, and Requirements.

2. Verification with Standard Block:
   Extract the block ID, standard value, the five individual readings, the average, the % error, and the remark.

3. Extracted Hardness Values Table:
   Create a table containing Sr. No., Sample ID No. (Pipe No,), and Heat No.
   For the 33 hardness measurement columns, group them into three sub-columns as labeled in the image: 
   Base (Points 1-6), HAZ (Points 7-24), and Weld (Points 25-33). 
   List the numbers for each row separated by commas within those groups. Include the 'Remarks' column at the end.

IMPORTANT: 
- Do NOT convert any numbers to decimals. Keep all numbers as integers.
- Do NOT add .0 to any numbers. For example, write "2601253" not "2601253.0".
- Heat No must be plain digits only, no decimal points.

Ensure all handwritten numbers are transcribed accurately and maintain a clean, professional layout.
"""

    client = genai.Client(api_key=st.secrets["GEMINI_API_KEY"])

    if st.session_state.extracted_text is None:
        with st.spinner("Extracting data from image..."):
            try:
                user_content = types.Content(
                    role="user",
                    parts=[
                        types.Part.from_bytes(mime_type="image/jpeg", data=image_bytes),
                        types.Part.from_text(text=prompt)
                    ]
                )

                response = client.models.generate_content(
                    model="gemini-3-flash-preview",  
                    contents=user_content,
                )
                st.session_state.extracted_text = response.text
                
                _, _, samples = parse_markdown_output(st.session_state.extracted_text)
                st.session_state.samples = samples
                
                st.success("✅ Data extracted successfully!")
                
            except Exception as e:
                st.error(f"An error occurred: {e}")
                st.stop()
    
    with st.expander("View Extracted Data"):
        st.code(st.session_state.extracted_text, language="markdown")
    
    if st.session_state.samples:
        st.subheader("📋 Extracted Test Samples")
        samples_df = pd.DataFrame([{
            "Pipe No": s["sample_id"],
            "Heat No": s["heat_no"],
            "Test Conducted": "Yes" if s["test_conducted"] else "No",
            "Base Readings": len(s["base"]),
            "HAZ Readings": len(s["haz"]),
            "Weld Readings": len(s["weld"])
        } for s in st.session_state.samples])
        st.dataframe(samples_df, use_container_width=True)
    
    if enable_validation and st.session_state.master_list_grouped is not None:
        st.subheader("📚 Master List (Heat No → Associated Pipe Nos)")
        st.caption(f"Total: {len(st.session_state.master_list_grouped)} unique Heat Numbers")
        
        st.dataframe(
            st.session_state.master_list_grouped,
            use_container_width=True,
            height=600,
            column_config={
                "Heat No": st.column_config.TextColumn("Heat No", width="small"),
                "Associated Pipe Nos": st.column_config.TextColumn("Associated Pipe Nos", width="large"),
                "Number of Pipes": st.column_config.NumberColumn("Number of Pipes", width="small")
            }
        )

        st.subheader("🔍 View All Pipe Numbers by Heat Number")
        for _, row in st.session_state.master_list_grouped.iterrows():
            with st.expander(f"Heat No: {row['Heat No']} ({row['Number of Pipes']} pipes)"):
                st.code(row['Associated Pipe Nos'], language="text")
        
   
    if enable_validation and st.session_state.master_map and st.session_state.samples and st.session_state.all_heats_in_master:
        st.subheader("Validation Results")
        
        results_df = validate_with_master_fast(st.session_state.samples, st.session_state.master_map, st.session_state.all_heats_in_master)
        
        if not results_df.empty:
            # Calculate correct counts
            total_heats = len(st.session_state.all_heats_in_master)
            tested_matched = len(results_df[results_df['Status'] == '✅ TESTED - MATCHED'])
            tested_mismatch = len(results_df[results_df['Status'] == '⚠️ TESTED - MISMATCH'])
            tested_not_in_master = len(results_df[results_df['Status'] == '❌ TESTED - NOT IN MASTER'])
            total_tested = tested_matched + tested_mismatch + tested_not_in_master
            not_tested = len(results_df[results_df['Status'] == '❌ NOT TESTED'])
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Unique Heat Nos", total_heats)
            with col2:
                st.metric("✅ Tested", total_tested)
            with col3:
                st.metric("❌ Not Tested", not_tested)
            
            st.dataframe(results_df, use_container_width=True)

    
    # Generate Excel files
    st.subheader("Download Reports")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("Generate Hardness Test Excel", type="primary", use_container_width=True):
            with st.spinner("Creating Hardness Test Excel file..."):
                TEMPLATE_PATH = "TEMPLATE MTC - Single Sheet.xlsx"
                OUTPUT_PATH = "Hardness_Test_Report.xlsx"
                
                try:
                    populate_excel_from_markdown(
                        st.session_state.extracted_text,
                        TEMPLATE_PATH,
                        OUTPUT_PATH
                    )
                    st.session_state.excel_ready = True
                    st.success("✅ Hardness Test Excel generated!")
                except Exception as e:
                    st.error(f"Error: {e}")
    
    with col2:
        if st.session_state.get('excel_ready', False):
            with open("Hardness_Test_Report.xlsx", "rb") as f:
                st.download_button(
                    "⬇️ Download Hardness Test Report",
                    data=f,
                    file_name="Vickers_Hardness_Test_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
    
    if enable_validation and st.session_state.master_map and st.session_state.samples and st.session_state.all_heats_in_master:
        if st.button("✅ Generate Validation Report", type="secondary", use_container_width=True):
            with st.spinner("Creating Validation Excel file..."):
                OUTPUT_PATH = "Validation_Report.xlsx"
                
                try:
                    create_validation_excel(
                        st.session_state.samples,
                        st.session_state.master_map,
                        st.session_state.all_heats_in_master,
                        OUTPUT_PATH
                    )
                    st.session_state.validation_excel_ready = True
                    st.success("✅ Validation Report generated!")
                except Exception as e:
                    st.error(f"Error: {e}")
        
        if st.session_state.get('validation_excel_ready', False):
            with open("Validation_Report.xlsx", "rb") as f:
                st.download_button(
                    "⬇️ Download Validation Report",
                    data=f,
                    file_name="Validation_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

# Clear cache button
st.sidebar.markdown("---")
if st.sidebar.button("🗑️ Clear Cache & Reset", use_container_width=True):
    st.session_state.extracted_text = None
    st.session_state.samples = None
    st.session_state.master_map = None
    st.session_state.master_list_grouped = None
    st.session_state.all_heats_in_master = None
    st.session_state.excel_ready = False
    st.session_state.validation_excel_ready = False
    st.rerun()

st.sidebar.markdown("---")
st.sidebar.info("""
**Instructions:**
1. Upload observation sheet image
2. Enable validation and upload master Excel
3. View Master List (heat numbers with associated pipe numbers)
4. Click 'Generate Hardness Test Excel'
5. Click 'Generate Validation Report'
6. Download both reports
""")