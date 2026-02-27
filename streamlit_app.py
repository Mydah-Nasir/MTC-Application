import streamlit as st
import base64
from google import genai
from google.genai import types
from PIL import Image
import io
import pandas as pd
from openpyxl import load_workbook
import re
from typing import List, Dict
# -------------------------
# Streamlit App
# -------------------------

def extract_bold_key_value(line: str):
    """
    Extracts **Key:** Value safely from a markdown line.
    Returns (key, value) or (None, None)
    """
    match = re.search(r"\*\*(.+?)\*\*:\s*(.+)", line)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    return None, None
def parse_markdown_output(md_text: str):

    summary = {}
    standard = {}
    rows = []

    # -----------------------------
    # 1. Test Summary
    # -----------------------------
    summary_section = re.search(
        r"### 1\. Test Summary Information(.*?)---",
        md_text,
        re.S
    )

    if summary_section:
        for line in summary_section.group(1).splitlines():
            key, val = extract_bold_key_value(line)
            if key:
                summary[key] = val

    # -----------------------------
    # 2. Standard Block
    # -----------------------------
    std_section = re.search(
        r"### 2\. Verification with Standard Block(.*?)---",
        md_text,
        re.S
    )

    if std_section:
        for line in std_section.group(1).splitlines():
            key, val = extract_bold_key_value(line)
            if key:
                standard[key] = val

    # -----------------------------
    # 3. Hardness Table
    # -----------------------------
    table_match = re.search(
        r"\| Sr\. No\..*",
        md_text,
        re.S
    )

    if table_match:
        lines = table_match.group(0).splitlines()

        # Skip header + separator
        data_lines = [
            l for l in lines
            if l.startswith("|") and not "---" in l
        ][1:]

        for line in data_lines:
            cols = [c.strip() for c in line.strip("|").split("|")]

            if len(cols) < 7:
                continue  # skip malformed rows safely

            rows.append({
                "sample_id": cols[1],
                "heat_no": cols[2],
                "base": [int(x.strip()) for x in cols[3].split(",") if x.strip().isdigit()],
                "haz": [int(x.strip()) for x in cols[4].split(",") if x.strip().isdigit()],
                "weld": [int(x.strip()) for x in cols[5].split(",") if x.strip().isdigit()],
                "remarks": cols[6]
            })

    return summary, standard, rows
def populate_excel_from_markdown(md_text, template_path, output_path):
    summary, standard, samples = parse_markdown_output(md_text)

    wb = load_workbook(template_path)
    ws = wb.active

    # -----------------------------
    # Map summary fields (EDIT CELLS ONCE)
    # -----------------------------
    summary_map = {
        "Testing Laboratory": "B5",
        "Document Title": "B6",
        "Format No.": "B7",
        "Specification & Grade": "B8",
        "Test Method": "B9",
        "Pipe Size": "B12",
        "Atmospheric Conditions": "B13",
        "Date & Shift": "B14",
        "Requirements": "B15",
        "M/C No.": "B16",
    }

    for k, cell in summary_map.items():
        if k in summary:
            ws[cell] = summary[k]

    # -----------------------------
    # Standard Block
    # -----------------------------
    std_map = {
        "Standard Block ID No.": "E5",
        "Standard Block Value": "E6",
        "Reading 1": "E7",
        "Reading 2": "E8",
        "Reading 3": "E9",
        "Reading 4": "E10",
        "Reading 5": "E11",
        "Average (AVG)": "E12",
        "% Of Error": "E13",
        "Remark": "E14",
    }

    for k, cell in std_map.items():
        if k in standard:
            ws[cell] = standard[k]

    # -----------------------------
    # Hardness Table
    # -----------------------------
    START_ROW = 17

    COL_SRNO   = 1   # Column A
    COL_SAMPLE = 2   # Column B
    COL_HEAT   = 3   # Column C
    COL_BASE   = 4   # Column D (Points 1–6)
    COL_HAZ    = 10  # Column J (Points 7–24)
    COL_WELD   = 28  # Column AB (Points 25–33)
    COL_REMARK = 37  # Column AK

    for idx, s in enumerate(samples):
        r = START_ROW + idx
        # Sr. No.
        ws.cell(r, COL_SRNO, idx + 1)

        # IDs
        ws.cell(r, COL_SAMPLE, s["sample_id"])
        ws.cell(r, COL_HEAT, s["heat_no"])

        # Base (1–6)
        for i, v in enumerate(s["base"]):
            ws.cell(r, COL_BASE + i, v)

        # HAZ (7–24)
        for i, v in enumerate(s["haz"]):
            ws.cell(r, COL_HAZ + i, v)

        # Weld (25–33)
        for i, v in enumerate(s["weld"]):
            ws.cell(r, COL_WELD + i, v)

        # Remarks
        ws.cell(r, COL_REMARK, s["remarks"])

    wb.save(output_path)
st.set_page_config(page_title="Vickers Hardness Sheet Extractor", layout="wide")
st.title("Vickers Hardness Observation Sheet Extractor")

st.write("""
Upload an 'Observation Sheet (Mechanical - Vickers Hardness Test)' image and extract structured Markdown output.
The output will include:
- Test Summary Information
- Verification with Standard Block
- Extracted Hardness Values Table
""")

# Upload image
uploaded_file = st.file_uploader("Upload Observation Sheet Image", type=["png", "jpg", "jpeg", "tiff"])

if uploaded_file:
    image = Image.open(uploaded_file)
    st.image(image, caption="Uploaded Observation Sheet", use_container_width=True)

    # Convert PIL image to bytes
    buffered = io.BytesIO()
    image.save(buffered, format="JPEG")
    image_bytes = buffered.getvalue()

    # Prompt for Gemini 3 Flash Preview
    prompt = """
Extract all information from the provided 'Observation Sheet (Mechanical - Vickers Hardness Test)' image into a structured Markdown format. 
Please organize the output into the following three sections:

1. Test Summary Information:
   Extract all metadata from the top section, including Customer, Spec & Grade, Test Method, Pipe Size, Atmospheric Conditions, Date, and Requirements.

2. Verification with Standard Block:
   Extract the block ID, standard value, the five individual readings, the average, the % error, and the remark.

3. Extracted Hardness Values Table:
   Create a table containing Sr. No., Sample ID No., and Heat No.
   For the 33 hardness measurement columns, group them into three sub-columns as labeled in the image: 
   Base (Points 1-6), HAZ (Points 7-24), and Weld (Points 25-33). 
   List the numbers for each row separated by commas within those groups. Include the 'Remarks' column at the end.

Ensure all handwritten numbers are transcribed accurately and maintain a clean, professional layout.
"""

    # Initialize Gemini client using st.secrets
    client = genai.Client(api_key=st.secrets["GEMINI_API_KEY"])

    with st.spinner("Extracting data from image..."):
        try:
            # Create user message with image and prompt
            user_content = types.Content(
                role="user",
                parts=[
                    types.Part.from_bytes(
                        mime_type="image/jpeg",
                        data=image_bytes
                    ),
                    types.Part.from_text(
                        text=prompt
                    )
                ]
            )


            response = client.models.generate_content(
                model="gemini-3-flash-preview",
                contents=user_content,
            )
            extracted_text = response.text
            print(extracted_text)
            st.subheader("Extracted Markdown Output")
            st.code(extracted_text, language="markdown")
            TEMPLATE_PATH = "TEMPLATE MTC - Single Sheet.xlsx"
            OUTPUT_PATH = "Populated_MTC.xlsx"

            populate_excel_from_markdown(
                extracted_text,
                TEMPLATE_PATH,
                OUTPUT_PATH
            )

            with open(OUTPUT_PATH, "rb") as f:
                st.download_button(
                    "Download Filled Excel File",
                    data=f,
                    file_name="Vickers_Hardness_MTC.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as e:
            print(f"Error: {e}")
            st.error(f"An error occurred: {e}")

